import time
import re
from bs4 import BeautifulSoup as bs
import requests as req
import openpyxl
import random
from user_collections.users import arr
from user_collections.CPMy import arr1
from var import key


proxies = [

]

testUser = [

]

toSave = [
    
]

sesh = req.Session()
baseURL = "https://google.com/search?q="

def populateArr():
    df = openpyxl.load_workbook("prospects.xlsx")
    df1 = df.active
    for row in range(1, df1.max_row):
        print(df1[row][6].value)
        toSave.append((df1[row][6].value, df1[row][7].value))
    with open("users", "w") as file:
        for item in toSave:
            file.write(str(item) + ",\n")

def saveToExcel():
    wb = openpyxl.Workbook()
    ws = wb.active
    for tuples in toSave:
        ws.append([tuples[0], tuples[1], tuples[2]])
    wb.save("trademark_usage.xlsx")

def executeSearch(fname, lname, abbr, cert):
    try: 
        #We could use a request session here I guess with rotating IPs
        payload = { 'api_key': key, 'url': f'{baseURL}{fname}+{lname}+{cert}'}
        response2 = req.get('https://api.scraperapi.com/', params=payload)
        print(f"{baseURL}{fname}+{lname}+{cert}")
    except Exception as e:
        print(e)
        return False
    print(response2.status_code)
    response2 = str(bs((response2.content), "html.parser"))
    match = re.search((fr"(?i)(?:{fname}\s*?.{{0,11}}?\s*?{lname}[\s\S]{{0,15}}?{abbr})"), response2)
    if (match):
        toSave.append([fname, lname, match.group()])
        return True
    else:
        return False

def main():
    count=0
    for identity in arr1:
        if (count <= 125):
            count+= 1
            print(identity)
            returner = executeSearch(identity[0], identity[1], "CPM", "Certified Property Manager")
            print(count, returner)
        else:
            saveToExcel()
            break

main()
